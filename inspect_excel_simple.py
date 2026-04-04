
from openpyxl import load_workbook
import os

file_path = r"c:\Users\HomePC\mdh_intranet\media\documents\2026\02\SimpleTabulation-ICD-11-MMS-en.xlsx"

try:
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(min_row=1, max_row=2, values_only=True)
    header = next(rows)
    first_row = next(rows)
    print("Headers:", header)
    print("First Row:", first_row)
except Exception as e:
    print(f"Error: {e}")
