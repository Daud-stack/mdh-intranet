from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.utils import timezone
from django.conf import settings
from django import forms
from .models import PreauthorizationRequest
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import json


# ── Forms ────────────────────────────────────────────────────────────────────

class RequestForm(forms.ModelForm):
    class Meta:
        model = PreauthorizationRequest
        fields = [
            'patient_id', 'patient_name', 'patient_dob', 'patient_gender', 'patient_phone',
            'scheme', 'scheme_plan', 'member_number', 'principal_name', 'relationship',
            'diagnosis', 'icd_code', 'procedure', 'procedure_code', 'clinical_notes', 'is_emergency',
            'referring_doctor', 'attending_doctor', 'facility_name',
            'amount', 'currency',
            'supporting_document',
        ]
        widgets = {
            'patient_id': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'e.g. PAT-001'}),
            'patient_name': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Full name as on medical aid card'}),
            'patient_dob': forms.DateInput(attrs={'class': 'form-control', 'type': 'date'}),
            'patient_gender': forms.Select(attrs={'class': 'form-select'}),
            'patient_phone': forms.TextInput(attrs={'class': 'form-control', 'placeholder': '+263 77 123 4567'}),
            'scheme': forms.Select(attrs={'class': 'form-select'}),
            'scheme_plan': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'e.g. Executive, Classic'}),
            'member_number': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Medical aid membership number'}),
            'principal_name': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Name of principal member'}),
            'relationship': forms.Select(attrs={'class': 'form-select'}),
            'diagnosis': forms.Textarea(attrs={'class': 'form-control', 'rows': 3, 'placeholder': 'Primary diagnosis or clinical indication'}),
            'icd_code': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'e.g. BA00'}),
            'procedure': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'e.g. MRI Brain, CT Scan Abdomen'}),
            'procedure_code': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'e.g. 70553'}),
            'clinical_notes': forms.Textarea(attrs={'class': 'form-control', 'rows': 3, 'placeholder': 'Additional clinical justification'}),
            'is_emergency': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'referring_doctor': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Dr. referring the patient'}),
            'attending_doctor': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Treating doctor/specialist'}),
            'facility_name': forms.TextInput(attrs={'class': 'form-control'}),
            'amount': forms.NumberInput(attrs={'class': 'form-control', 'placeholder': '0.00', 'step': '0.01'}),
            'currency': forms.Select(attrs={'class': 'form-select'}),
            'supporting_document': forms.FileInput(attrs={'class': 'form-control'}),
        }


class EmailSendForm(forms.Form):
    """Form for sending preauth to medical aid via email."""
    from_email = forms.EmailField(
        label='From (Your Email)',
        widget=forms.EmailInput(attrs={
            'class': 'form-control',
            'placeholder': 'your-email@example.com',
            'id': 'id_from_email',
        })
    )
    to_email = forms.EmailField(
        label='To (Medical Aid)',
        widget=forms.EmailInput(attrs={
            'class': 'form-control',
            'placeholder': 'preauth@medicalaid.co.zw',
            'id': 'id_to_email',
        })
    )
    cc_email = forms.EmailField(
        label='CC (optional)',
        required=False,
        widget=forms.EmailInput(attrs={
            'class': 'form-control',
            'placeholder': 'Optional CC address',
        })
    )
    subject = forms.CharField(
        label='Email Subject',
        widget=forms.TextInput(attrs={
            'class': 'form-control',
        })
    )
    message = forms.CharField(
        label='Additional Message',
        required=False,
        widget=forms.Textarea(attrs={
            'class': 'form-control',
            'rows': 4,
            'placeholder': 'Optional message to include in the email body',
        })
    )


# ── Views ────────────────────────────────────────────────────────────────────

@login_required
def create_request(request):
    if request.method == 'POST':
        form = RequestForm(request.POST, request.FILES)
        if form.is_valid():
            req = form.save(commit=False)
            req.submitted_by = request.user
            req.save()
            messages.success(request, 'Preauthorization request submitted! You can now send it to the medical aid.')
            return redirect('medical_aid:detail', pk=req.pk)
    else:
        form = RequestForm()
    return render(request, 'medical_aid/create.html', {'form': form})


@login_required
def request_detail(request, pk):
    req = get_object_or_404(PreauthorizationRequest, pk=pk)
    email_form = EmailSendForm(initial={
        'from_email': request.user.email or settings.EMAIL_HOST_USER,
        'to_email': req.sent_to_email or req.default_scheme_email,
        'subject': f'Preauthorization Request - {req.patient_name} ({req.patient_id}) - Ref PA-{req.pk:04d}',
    })
    scheme_emails_json = json.dumps(PreauthorizationRequest.SCHEME_EMAILS)
    return render(request, 'medical_aid/detail.html', {
        'req': req,
        'email_form': email_form,
        'scheme_emails_json': scheme_emails_json,
    })


@login_required
def request_list(request):
    requests = PreauthorizationRequest.objects.all()
    context = {
        'requests': requests,
        'pending_count': requests.filter(status='PENDING').count(),
        'approved_count': requests.filter(status='APPROVED').count(),
        'rejected_count': requests.filter(status='REJECTED').count(),
    }
    return render(request, 'medical_aid/list.html', context)


# ── Excel Generation ─────────────────────────────────────────────────────────

def _generate_excel(req):
    """Generate the preauth Excel workbook and return (BytesIO buffer, filename)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Preauthorization Request'

    header_font = Font(name='Calibri', bold=True, size=16, color='FFFFFF')
    header_fill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
    section_font = Font(name='Calibri', bold=True, size=11, color='1B4F72')
    section_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
    label_font = Font(name='Calibri', bold=True, size=10, color='2C3E50')
    value_font = Font(name='Calibri', size=10)
    thin_border = Border(
        left=Side(style='thin', color='BDC3C7'),
        right=Side(style='thin', color='BDC3C7'),
        top=Side(style='thin', color='BDC3C7'),
        bottom=Side(style='thin', color='BDC3C7'),
    )

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 35

    # Header
    ws.merge_cells('A1:D1')
    ws['A1'].value = 'HOSPITAL PREAUTHORIZATION'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    for col in ['B1', 'C1', 'D1']:
        ws[col].fill = header_fill

    ws.merge_cells('A2:D2')
    ws['A2'].value = 'MEDICAL AID PREAUTHORIZATION REQUEST FORM'
    ws['A2'].font = Font(name='Calibri', bold=True, size=12, color='1B4F72')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 28

    ws.merge_cells('A3:D3')
    ws['A3'].value = f'Date: {req.created_at.strftime("%d %B %Y")}    |    Reference: PA-{req.pk:04d}    |    Status: {req.get_status_display()}'
    ws['A3'].font = Font(name='Calibri', size=9, italic=True, color='7F8C8D')
    ws['A3'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 8

    def write_section(r, title):
        ws.merge_cells(f'A{r}:D{r}')
        ws[f'A{r}'].value = title
        ws[f'A{r}'].font = section_font
        ws[f'A{r}'].fill = section_fill
        ws[f'A{r}'].alignment = Alignment(vertical='center')
        ws.row_dimensions[r].height = 26
        for c in ['B', 'C', 'D']:
            ws[f'{c}{r}'].fill = section_fill
        return r + 1

    def write_pair(r, l1, v1, l2=None, v2=None):
        ws[f'A{r}'].value = l1
        ws[f'A{r}'].font = label_font
        ws[f'A{r}'].border = thin_border
        ws[f'A{r}'].alignment = Alignment(vertical='center')
        ws[f'B{r}'].value = str(v1) if v1 else ''
        ws[f'B{r}'].font = value_font
        ws[f'B{r}'].border = thin_border
        ws[f'B{r}'].alignment = Alignment(vertical='center')
        if l2:
            ws[f'C{r}'].value = l2
            ws[f'C{r}'].font = label_font
            ws[f'C{r}'].border = thin_border
            ws[f'C{r}'].alignment = Alignment(vertical='center')
            ws[f'D{r}'].value = str(v2) if v2 else ''
            ws[f'D{r}'].font = value_font
            ws[f'D{r}'].border = thin_border
            ws[f'D{r}'].alignment = Alignment(vertical='center')
        else:
            ws.merge_cells(f'B{r}:D{r}')
        ws.row_dimensions[r].height = 22
        return r + 1

    def write_full(r, label, value):
        ws[f'A{r}'].value = label
        ws[f'A{r}'].font = label_font
        ws[f'A{r}'].border = thin_border
        ws[f'A{r}'].alignment = Alignment(vertical='top')
        ws.merge_cells(f'B{r}:D{r}')
        ws[f'B{r}'].value = str(value) if value else ''
        ws[f'B{r}'].font = value_font
        ws[f'B{r}'].border = thin_border
        ws[f'B{r}'].alignment = Alignment(vertical='top', wrap_text=True)
        ws.row_dimensions[r].height = max(22, len(str(value or '')) // 60 * 15 + 22)
        return r + 1

    row = 5
    row = write_section(row, '1. PATIENT INFORMATION')
    row = write_pair(row, 'Patient / Member ID:', req.patient_id, 'Full Name:', req.patient_name)
    row = write_pair(row, 'Date of Birth:', req.patient_dob.strftime('%d/%m/%Y') if req.patient_dob else '', 'Gender:', req.patient_gender)
    row = write_pair(row, 'Contact Number:', req.patient_phone, 'Emergency:', 'YES' if req.is_emergency else 'No')

    row += 1
    row = write_section(row, '2. MEDICAL AID / INSURANCE DETAILS')
    row = write_pair(row, 'Scheme:', req.get_scheme_display(), 'Plan / Tier:', req.scheme_plan)
    row = write_pair(row, 'Member Number:', req.member_number, 'Relationship:', req.relationship)
    row = write_pair(row, 'Principal Member:', req.principal_name, '', '')

    row += 1
    row = write_section(row, '3. CLINICAL INFORMATION')
    row = write_full(row, 'Diagnosis:', req.diagnosis)
    row = write_pair(row, 'ICD-11 Code:', req.icd_code, 'Procedure Code (CPT):', req.procedure_code)
    row = write_full(row, 'Procedure / Treatment:', req.procedure)
    row = write_full(row, 'Clinical Notes:', req.clinical_notes)

    row += 1
    row = write_section(row, '4. PROVIDER DETAILS')
    row = write_pair(row, 'Facility:', req.facility_name, 'Referring Doctor:', req.referring_doctor)
    row = write_pair(row, 'Attending Doctor:', req.attending_doctor, '', '')

    row += 1
    row = write_section(row, '5. FINANCIAL INFORMATION')
    amount_str = f'{req.currency} {req.amount:,.2f}' if req.amount else ''
    row = write_pair(row, 'Estimated Cost:', amount_str, 'Auth Number:', req.auth_number)

    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].value = 'FOR MEDICAL AID USE ONLY'
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    for c in ['B', 'C', 'D']:
        ws[f'{c}{row}'].fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
    row += 1
    row = write_pair(row, 'Authorization Number:', '', 'Date Authorized:', '')
    row = write_pair(row, 'Authorized By:', '', 'Approved Amount:', '')
    row = write_full(row, 'Comments:', '')

    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].value = f'Generated by OpsHub on {req.created_at.strftime("%d %B %Y at %H:%M")}  •  System-generated document.'
    ws[f'A{row}'].font = Font(name='Calibri', size=8, italic=True, color='95A5A6')
    ws[f'A{row}'].alignment = Alignment(horizontal='center')

    ws.print_area = f'A1:D{row}'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, f'Preauth_{req.patient_id}_{req.pk}.xlsx'


# ── Download Excel ───────────────────────────────────────────────────────────

@login_required
def export_excel(request, pk):
    """Download the preauth form as an Excel file."""
    req = get_object_or_404(PreauthorizationRequest, pk=pk)
    buffer, filename = _generate_excel(req)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── Send Email ───────────────────────────────────────────────────────────────

@login_required
def send_email(request, pk):
    """Send the preauth Excel form to the medical aid via email."""
    req = get_object_or_404(PreauthorizationRequest, pk=pk)

    if request.method != 'POST':
        return redirect('medical_aid:detail', pk=pk)

    form = EmailSendForm(request.POST)
    if not form.is_valid():
        messages.error(request, 'Please fill in a valid email address.')
        return redirect('medical_aid:detail', pk=pk)

    from_email = form.cleaned_data['from_email']
    to_email = form.cleaned_data['to_email']
    cc_email = form.cleaned_data.get('cc_email', '')
    subject = form.cleaned_data['subject']
    extra_message = form.cleaned_data.get('message', '')

    # Generate Excel
    buffer, filename = _generate_excel(req)

    # Build email body
    body = f"""Dear {req.get_scheme_display()} Preauthorization Team,

Please find attached a preauthorization request from the hospital.

REQUEST SUMMARY
{'-' * 50}
Reference:      PA-{req.pk:04d}
Patient:        {req.patient_name} ({req.patient_id})
Member No:      {req.member_number or 'N/A'}
Diagnosis:      {req.diagnosis}
Procedure:      {req.procedure}
Estimated Cost: {f'{req.currency} {req.amount:,.2f}' if req.amount else 'N/A'}
Emergency:      {'YES' if req.is_emergency else 'No'}
{'-' * 50}
"""
    if extra_message:
        body += f"\nAdditional Notes:\n{extra_message}\n"

    sender_name = request.user.get_full_name() or request.user.username
    body += f"""
Please process this request and provide an authorization number at your earliest convenience.

Thank you,
{sender_name}
Hospital Administration
"""

    try:
        email = EmailMessage(
            subject=subject,
            body=body,
            from_email=from_email,
            to=[to_email],
            cc=[cc_email] if cc_email else [],
            reply_to=[from_email],
        )
        email.attach(filename, buffer.getvalue(),
                     'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Also attach supporting document if present
        if req.supporting_document:
            try:
                email.attach_file(req.supporting_document.path)
            except Exception:
                pass

        email.send(fail_silently=False)

        # Track on the model
        req.sent_to_email = to_email
        req.email_sent_at = timezone.now()
        req.email_sent_by = request.user
        req.save(update_fields=['sent_to_email', 'email_sent_at', 'email_sent_by'])

        messages.success(request, f'Preauthorization form sent successfully to {to_email}!')

    except Exception as e:
        messages.error(request, f'Failed to send email: {str(e)}. Please check your email settings.')

    return redirect('medical_aid:detail', pk=pk)
